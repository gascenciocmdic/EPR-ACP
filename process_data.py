import openpyxl
import json
import datetime
import os

def clean_val(val):
    if val is None:
        return None
    if isinstance(val, str):
        val_stripped = val.strip()
        if val_stripped == "" or val_stripped == "-":
            return None
        return val_stripped
    return val

def clean_num(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val_str = val.strip().replace(",", ".")
        if val_str == "" or val_str == "-":
            return 0.0
        try:
            return float(val_str)
        except ValueError:
            return 0.0
    return 0.0

def get_clean_discipline(col_val, section_val):
    val = clean_val(col_val)
    if val and val != "DISCIPLINA":
        return val
        
    if not section_val:
        return "General / Gestión"
        
    s = section_val.lower()
    if "arqueología" in s:
        return "Arqueología"
    if "animales" in s or "fauna" in s or "invertebrados" in s or "invertebrada" in s:
        return "Fauna y Biodiversidad"
    if "plantas" in s or "vegetación" in s or "hongos" in s or "líquenes" in s or "vegetal" in s or "metharme" in s:
        return "Flora y Vegetación"
    if "aguas" in s or "hidro" in s or "pit-lake" in s or "infiltración" in s or "sumidero" in s or "modelo" in s:
        return "Hidrología e Hidrogeología"
    if "suelos" in s:
        return "Suelos"
    if "ruido" in s or "vibración" in s or "vibraciones" in s:
        return "Ruido y Vibraciones"
    if "antropológico" in s or "grupos humanos" in s or "humano" in s or "pac" in s:
        return "Medio Humano / Social"
    if "elaboración" in s or "capitulos" in s or "eia" in s or "anexos" in s:
        return "Elaboración EIA"
    if "vial" in s:
        return "Vial / Transporte"
    if "paisaje" in s:
        return "Paisaje"
    if "paleontología" in s:
        return "Paleontología"
    if "campaña" in s or "preparación" in s or "estación" in s:
        return "Campañas de Terreno"
    return "General / Gestión"

def clean_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val_str = val.strip()
        if val_str == "" or val_str == "-":
            return None
        # If it has newlines, take the last non-empty line
        lines = [l.strip() for l in val_str.split("\n") if l.strip()]
        if lines:
            val_str = lines[-1]
        # Only strip by space if it matches a timestamp pattern
        import re
        if re.match(r"^\d{4}-\d{2}-\d{2}\s", val_str):
            val_str = val_str.split(" ")[0]
        elif re.match(r"^\d{2}[-/]\d{2}[-/]\d{2,4}\s", val_str):
            val_str = val_str.split(" ")[0]
        return val_str
    return str(val)

def parse_date_obj(date_str):
    if not date_str:
        return None
    # Try parsing YYYY-MM-DD
    try:
        return datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        pass
    # Try parsing DD/MM/YY
    try:
        # e.g. "06/06/25" or "06/06/2025"
        parts = date_str.split("/")
        if len(parts) == 3:
            day = int(parts[0])
            month = int(parts[1])
            year = int(parts[2])
            if year < 100:
                year += 2000
            return datetime.date(year, month, day)
    except ValueError:
        pass
    # Try parsing DD-MM-YYYY
    try:
        parts = date_str.split("-")
        if len(parts) == 3:
            # check if it is YYYY-MM-DD or DD-MM-YYYY
            if len(parts[0]) == 4:
                return datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
            else:
                day = int(parts[0])
                month = int(parts[1])
                year = int(parts[2])
                if year < 100:
                    year += 2000
                return datetime.date(year, month, day)
    except ValueError:
        pass
    
    # Try to extract any date-like string from text like "P1 06/06/25"
    import re
    date_match = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", date_str)
    if date_match:
        try:
            day = int(date_match.group(1))
            month = int(date_match.group(2))
            year = int(date_match.group(3))
            if year < 100:
                year += 2000
            return datetime.date(year, month, day)
        except ValueError:
            pass
            
    return None

def extract_epr_data(filepath):
    print(f"Extracting {filepath}...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb["EPR"]
    
    data_rows = []
    current_section = "General"
    
    for r in range(11, sheet.max_row + 1):
        c_val = clean_val(sheet.cell(row=r, column=3).value) # Col C (Activity ID)
        f_val = clean_val(sheet.cell(row=r, column=6).value) # Col F (Description)
        
        # Check if this row is the end note
        if c_val and len(c_val) > 50 and "segregación" in c_val:
            break
            
        # Check if this is a section header (No C, but has F)
        if c_val is None and f_val is not None:
            # Check if it has HH values, if so, it might be a summary row, not just a section
            hh_val = sheet.cell(row=r, column=13).value
            if hh_val is None or hh_val == 0:
                current_section = f_val
                continue
            else:
                # It's a summary row, skip it as we will aggregate manually
                continue
                
        if c_val is None and f_val is None:
            continue
            
        # Extract fields
        item = {
            "row": r,
            "section": current_section,
            "activity_id": c_val,
            "status": clean_val(sheet.cell(row=r, column=4).value), # Col D
            "cmdic_code": clean_val(sheet.cell(row=r, column=5).value), # Col E
            "description": f_val,
            "discipline": get_clean_discipline(sheet.cell(row=r, column=7).value, current_section), # Col G
            "is_document": True if clean_val(sheet.cell(row=r, column=8).value) is not None else False, # Col H
            "is_plano": True if clean_val(sheet.cell(row=r, column=9).value) is not None else False, # Col I
            "is_activity": True if clean_val(sheet.cell(row=r, column=10).value) is not None else False, # Col J
            "hh_co": clean_num(sheet.cell(row=r, column=11).value),
            "hh_td": clean_num(sheet.cell(row=r, column=12).value),
            "hh_total_ppto": clean_num(sheet.cell(row=r, column=13).value),
            "hh_total_ganadas": clean_num(sheet.cell(row=r, column=14).value),
            "avance_programado": clean_num(sheet.cell(row=r, column=15).value),
            "avance_real": clean_num(sheet.cell(row=r, column=16).value),
        }
        
        # Determine item type
        if item["is_document"]:
            item["type"] = "Documento"
        elif item["is_plano"]:
            item["type"] = "Plano"
        else:
            item["type"] = "Actividad"
            
        # PLAN dates (Q to V)
        item["plan"] = {
            "inicio": clean_date(sheet.cell(row=r, column=17).value),
            "emision_reva": clean_date(sheet.cell(row=r, column=18).value),
            "emision_revb": clean_date(sheet.cell(row=r, column=19).value),
            "aprobacion_revb": clean_date(sheet.cell(row=r, column=20).value),
            "emision_revp": clean_date(sheet.cell(row=r, column=21).value),
            "aprobacion_revp": clean_date(sheet.cell(row=r, column=22).value)
        }
        
        # REAL dates (W to AD)
        item["real"] = {
            "inicio": clean_date(sheet.cell(row=r, column=23).value),
            "emision_reva": clean_date(sheet.cell(row=r, column=24).value),
            "emision_revb": clean_date(sheet.cell(row=r, column=25).value),
            "aprobacion_revb": clean_date(sheet.cell(row=r, column=26).value),
            "emision_revc": clean_date(sheet.cell(row=r, column=27).value),
            "aprobacion_revc": clean_date(sheet.cell(row=r, column=28).value),
            "emision_revp": clean_date(sheet.cell(row=r, column=29).value),
            "aprobacion_revp": clean_date(sheet.cell(row=r, column=30).value)
        }
        
        # FORECAST dates (AE to AJ)
        item["forecast"] = {
            "inicio": clean_date(sheet.cell(row=r, column=31).value),
            "emision_reva": clean_date(sheet.cell(row=r, column=32).value),
            "emision_revb": clean_date(sheet.cell(row=r, column=33).value),
            "aprobacion_revb": clean_date(sheet.cell(row=r, column=34).value),
            "emision_revp": clean_date(sheet.cell(row=r, column=35).value),
            "aprobacion_revp": clean_date(sheet.cell(row=r, column=36).value)
        }
        
        data_rows.append(item)
        
    print(f"  Successfully extracted {len(data_rows)} items.")
    return data_rows

def run_calculations(items):
    # Overall sum indicators
    total_hh_ppto = sum(x["hh_total_ppto"] for x in items)
    total_hh_ganadas = sum(x["hh_total_ganadas"] for x in items)
    total_hh_co = sum(x["hh_co"] for x in items)
    
    # Global progress weighted by HH Total Ppto
    global_avance_programado = sum(x["avance_programado"] * x["hh_total_ppto"] for x in items) / total_hh_ppto if total_hh_ppto > 0 else 0
    global_avance_real = sum(x["avance_real"] * x["hh_total_ppto"] for x in items) / total_hh_ppto if total_hh_ppto > 0 else 0
    
    # CPI (Cost Performance Index or rather, HH Efficiency Index = Earned / Planned-To-Date)
    # If planned progress is P, then Planned HH to date is P * total_hh_ppto
    planned_hh_to_date = sum(x["avance_programado"] * x["hh_total_ppto"] for x in items)
    hh_efficiency = total_hh_ganadas / planned_hh_to_date if planned_hh_to_date > 0 else 1.0
    
    # Count by status
    # Standard statuses:
    # Completed (avance_real = 1.0)
    # In Progress (0 < avance_real < 1.0)
    # Not Started (avance_real = 0.0)
    status_counts = {"completado": 0, "proceso": 0, "no_iniciado": 0}
    for x in items:
        if x["avance_real"] >= 1.0:
            status_counts["completado"] += 1
        elif x["avance_real"] > 0:
            status_counts["proceso"] += 1
        else:
            status_counts["no_iniciado"] += 1
            
    # Count by type
    type_counts = {"Documento": 0, "Plano": 0, "Actividad": 0}
    for x in items:
        type_counts[x["type"]] += 1
        
    # Analyze disciplines
    disciplines = {}
    for x in items:
        disc = x["discipline"]
        if disc not in disciplines:
            disciplines[disc] = {
                "count": 0,
                "hh_ppto": 0,
                "hh_ganadas": 0,
                "avance_prog_weighted_sum": 0,
                "avance_real_weighted_sum": 0,
                "completed": 0
            }
        d = disciplines[disc]
        d["count"] += 1
        d["hh_ppto"] += x["hh_total_ppto"]
        d["hh_ganadas"] += x["hh_total_ganadas"]
        d["avance_prog_weighted_sum"] += x["avance_programado"] * x["hh_total_ppto"]
        d["avance_real_weighted_sum"] += x["avance_real"] * x["hh_total_ppto"]
        if x["avance_real"] >= 1.0:
            d["completed"] += 1
            
    disciplines_summary = []
    for name, d in disciplines.items():
        hh_ppto = d["hh_ppto"]
        avg_prog = d["avance_prog_weighted_sum"] / hh_ppto if hh_ppto > 0 else 0
        avg_real = d["avance_real_weighted_sum"] / hh_ppto if hh_ppto > 0 else 0
        disciplines_summary.append({
            "name": name,
            "count": d["count"],
            "completed": d["completed"],
            "hh_ppto": hh_ppto,
            "hh_ganadas": d["hh_ganadas"],
            "avance_programado": avg_prog,
            "avance_real": avg_real,
            "desviacion": avg_real - avg_prog
        })
    # Sort disciplines by HH size
    disciplines_summary.sort(key=lambda x: x["hh_ppto"], reverse=True)
    
    return {
        "total_items": len(items),
        "total_hh_ppto": total_hh_ppto,
        "total_hh_ganadas": total_hh_ganadas,
        "total_hh_co": total_hh_co,
        "avance_programado": global_avance_programado,
        "avance_real": global_avance_real,
        "hh_efficiency": hh_efficiency,
        "status_counts": status_counts,
        "type_counts": type_counts,
        "disciplines": disciplines_summary
    }

def cross_reference(actual_items, pasada_items):
    pasada_dict = {x["activity_id"]: x for x in pasada_items if x["activity_id"]}
    
    cross_logs = []
    new_items = []
    deleted_items = []
    
    for item in actual_items:
        act_id = item["activity_id"]
        if not act_id:
            continue
            
        if act_id not in pasada_dict:
            new_items.append({
                "activity_id": act_id,
                "cmdic_code": item["cmdic_code"],
                "description": item["description"],
                "discipline": item["discipline"],
                "type": item["type"],
                "hh_ppto": item["hh_total_ppto"],
                "avance_real": item["avance_real"]
            })
            continue
            
        pasada_item = pasada_dict[act_id]
        
        # Compare progress
        progress_diff = item["avance_real"] - pasada_item["avance_real"]
        
        # Compare forecast finish date
        actual_forecast_end_str = item["forecast"]["aprobacion_revp"]
        pasada_forecast_end_str = pasada_item["forecast"]["aprobacion_revp"]
        
        actual_forecast_end = parse_date_obj(actual_forecast_end_str)
        pasada_forecast_end = parse_date_obj(pasada_forecast_end_str)
        
        days_slipped_this_week = 0
        if actual_forecast_end and pasada_forecast_end:
            days_slipped_this_week = (actual_forecast_end - pasada_forecast_end).days
        elif actual_forecast_end and not pasada_forecast_end:
            # Became scheduled
            days_slipped_this_week = 0
            
        # Compare plan finish date to forecast finish date (cumulative slippage)
        plan_end = parse_date_obj(item["plan"]["aprobacion_revp"])
        cumulative_slippage = 0
        if actual_forecast_end and plan_end:
            cumulative_slippage = (actual_forecast_end - plan_end).days
            
        # Check if status transitioned to completed
        status_transition = None
        if pasada_item["avance_real"] < 1.0 and item["avance_real"] >= 1.0:
            status_transition = "completado"
        elif pasada_item["avance_real"] == 0.0 and item["avance_real"] > 0.0:
            status_transition = "iniciado"
            
        # Check if actual start occurred this week
        started_this_week = False
        if not pasada_item["real"]["inicio"] and item["real"]["inicio"]:
            started_this_week = True
            
        if progress_diff > 0 or days_slipped_this_week != 0 or status_transition or started_this_week:
            cross_logs.append({
                "activity_id": act_id,
                "cmdic_code": item["cmdic_code"],
                "description": item["description"],
                "discipline": item["discipline"],
                "type": item["type"],
                "hh_ppto": item["hh_total_ppto"],
                "old_avance_real": pasada_item["avance_real"],
                "new_avance_real": item["avance_real"],
                "avance_real_diff": progress_diff,
                "old_forecast_end": pasada_forecast_end_str,
                "new_forecast_end": actual_forecast_end_str,
                "days_slipped_this_week": days_slipped_this_week,
                "cumulative_slippage": cumulative_slippage,
                "status_transition": status_transition,
                "started_this_week": started_this_week
            })
            
    # Check for deleted items
    actual_ids = {x["activity_id"] for x in actual_items if x["activity_id"]}
    for act_id, item in pasada_dict.items():
        if act_id not in actual_ids:
            deleted_items.append({
                "activity_id": act_id,
                "cmdic_code": item["cmdic_code"],
                "description": item["description"],
                "discipline": item["discipline"],
                "type": item["type"],
                "hh_ppto": item["hh_total_ppto"]
            })
            
    return {
        "changes": cross_logs,
        "new_items": new_items,
        "deleted_items": deleted_items
    }

def generate_s_curve_data(items):
    # To construct S-curves, we collect dates and planned/forecast/earned increments
    # Let's use planned start and planned end to distribute planned progress linearly.
    # Similarly, forecast start and forecast end to distribute forecast progress.
    # For actual earned progress, we know the actual earned hours today.
    # What about historical actual earned hours? We can look at completed tasks and their actual finish date,
    # and for in-progress tasks, we can assume they progressed linearly up to today (using actual start and today).
    # Since today's date is roughly May 24, 2026, let's use that as the "status date" or the max date.
    
    status_date = datetime.date(2026, 5, 24)
    
    # We will gather all dates from 2025-01-01 to 2027-12-31 and aggregate by month
    # Let's create monthly buckets
    start_year = 2025
    end_year = 2027
    
    months = []
    for y in range(start_year, end_year + 1):
        for m in range(1, 13):
            # We want month-end dates
            if m == 12:
                d = datetime.date(y, 12, 31)
            else:
                d = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)
            months.append(d)
            
    months.sort()
    
    # Let's distribute planned HH for each activity
    # For each month-end date, we calculate the cumulative planned progress
    planned_cum_hh = [0.0] * len(months)
    forecast_cum_hh = [0.0] * len(months)
    actual_cum_hh = [0.0] * len(months)
    
    total_hh = sum(x["hh_total_ppto"] for x in items)
    if total_hh == 0:
        total_hh = 1.0
        
    for x in items:
        hh = x["hh_total_ppto"]
        if hh <= 0:
            continue
            
        # 1. Planned Progress
        p_start = parse_date_obj(x["plan"]["inicio"])
        p_end = parse_date_obj(x["plan"]["aprobacion_revp"])
        
        # If no dates are set, default to a sensible fallback or ignore
        if p_start and p_end:
            p_dur = (p_end - p_start).days
            if p_dur <= 0:
                p_dur = 1
            for idx, m_date in enumerate(months):
                if m_date < p_start:
                    pct = 0.0
                elif m_date >= p_end:
                    pct = 1.0
                else:
                    pct = (m_date - p_start).days / p_dur
                planned_cum_hh[idx] += pct * hh
        else:
            # Fallback: if there is no plan start/end, but it is completed, assume it was planned by its actual approval or today
            p_end_alt = p_end or parse_date_obj(x["plan"]["aprobacion_revp"]) or status_date
            for idx, m_date in enumerate(months):
                if m_date >= p_end_alt:
                    planned_cum_hh[idx] += hh
                    
        # 2. Forecast Progress
        # Forecast dates default to Plan dates if not present
        f_start = parse_date_obj(x["forecast"]["inicio"]) or p_start
        f_end = parse_date_obj(x["forecast"]["aprobacion_revp"]) or p_end
        
        if f_start and f_end:
            f_dur = (f_end - f_start).days
            if f_dur <= 0:
                f_dur = 1
            for idx, m_date in enumerate(months):
                if m_date < f_start:
                    pct = 0.0
                elif m_date >= f_end:
                    pct = 1.0
                else:
                    pct = (m_date - f_start).days / f_dur
                forecast_cum_hh[idx] += pct * hh
        else:
            f_end_alt = f_end or status_date
            for idx, m_date in enumerate(months):
                if m_date >= f_end_alt:
                    forecast_cum_hh[idx] += hh
                    
        # 3. Actual Progress (Earned Hours)
        # For completed items, they earned their full HH on the actual approval date.
        # For in-progress items, they have earned a fraction of their HH as of today.
        # We want to reconstruct the curve of earned hours over time.
        # Today's actual earned hours is x["hh_total_ganadas"] or x["avance_real"] * hh.
        # How did we get here?
        # If the task is completed: it has an actual start (real.inicio) and actual end (real.aprobacion_revp).
        # We can assume it progressed linearly between actual start and actual end.
        # If actual end is not available but progress is 100%, we use status_date.
        # If the task is in progress: it has actual start (real.inicio) and current progress.
        # We can assume it progressed linearly from actual start to status_date.
        r_start = parse_date_obj(x["real"]["inicio"]) or p_start
        r_end = parse_date_obj(x["real"]["aprobacion_revp"])
        
        # Current earned
        earned_hh_now = x["hh_total_ganadas"]
        if earned_hh_now == 0 and x["avance_real"] > 0:
            earned_hh_now = x["avance_real"] * hh
            
        if earned_hh_now > 0:
            if x["avance_real"] >= 1.0 and r_start and r_end:
                # Completed task with dates
                r_dur = (r_end - r_start).days
                if r_dur <= 0:
                    r_dur = 1
                for idx, m_date in enumerate(months):
                    if m_date < r_start:
                        pct = 0.0
                    elif m_date >= r_end:
                        pct = 1.0
                    else:
                        pct = (m_date - r_start).days / r_dur
                    actual_cum_hh[idx] += pct * earned_hh_now
            elif r_start:
                # In progress or completed without end date
                # Progress goes from r_start to status_date
                r_dur = (status_date - r_start).days
                if r_dur <= 0:
                    r_dur = 1
                for idx, m_date in enumerate(months):
                    if m_date < r_start:
                        pct = 0.0
                    elif m_date >= status_date:
                        pct = 1.0
                    else:
                        pct = (m_date - r_start).days / r_dur
                    # We limit the earned progress to what is earned up to that month
                    actual_cum_hh[idx] += min(pct, 1.0) * earned_hh_now
            else:
                # Fallback: earned hours only appear at status_date
                for idx, m_date in enumerate(months):
                    if m_date >= status_date:
                        actual_cum_hh[idx] += earned_hh_now
                        
    # Convert cumulative HH to percentages
    planned_cum_pct = [round((val / total_hh) * 100, 2) for val in planned_cum_hh]
    forecast_cum_pct = [round((val / total_hh) * 100, 2) for val in forecast_cum_hh]
    actual_cum_pct = [round((val / total_hh) * 100, 2) for val in actual_cum_hh]
    
    # Trim the actual curve after status_date (since we cannot have actual progress in the future!)
    for idx, m_date in enumerate(months):
        if m_date > status_date:
            actual_cum_pct[idx] = None
            
    # Filter months to show only those where there is some activity (e.g. from April 2025 to Dec 2027)
    # Let's find first index where planned or actual progress > 0
    first_active_idx = 0
    for idx in range(len(months)):
        if planned_cum_pct[idx] > 0 or (actual_cum_pct[idx] is not None and actual_cum_pct[idx] > 0):
            first_active_idx = max(0, idx - 1) # include one month before start
            break
            
    # Find last index with data
    last_active_idx = len(months) - 1
    for idx in range(len(months) - 1, -1, -1):
        if planned_cum_pct[idx] < 100 or forecast_cum_pct[idx] < 100:
            last_active_idx = min(len(months) - 1, idx + 2) # include month when we hit 100%
            break
            
    slice_start = first_active_idx
    slice_end = last_active_idx + 1
    
    return {
        "labels": [m.strftime("%Y-%m") for m in months[slice_start:slice_end]],
        "planned": planned_cum_pct[slice_start:slice_end],
        "forecast": forecast_cum_pct[slice_start:slice_end],
        "actual": actual_cum_pct[slice_start:slice_end]
    }

def main():
    # Paths
    actual_xlsx = "Semana Actual.xlsx"
    pasada_xlsx = "Semana Pasada.xlsx"
    
    print("Starting processing...")
    actual_items = extract_epr_data(actual_xlsx)
    pasada_items = extract_epr_data(pasada_xlsx)
    
    # 1. Run Calculations
    print("Running calculations...")
    actual_calcs = run_calculations(actual_items)
    pasada_calcs = run_calculations(pasada_items)
    
    # 2. Run Cross-Reference Comparison
    print("Running cross-reference...")
    comparison = cross_reference(actual_items, pasada_items)
    
    # 3. Generate S-Curve Data
    print("Generating S-curve...")
    s_curve = generate_s_curve_data(actual_items)
    
    # 4. Find most delayed items (highest slippage)
    delayed_items = []
    for x in actual_items:
        if not x["activity_id"]:
            continue
        p_end = parse_date_obj(x["plan"]["aprobacion_revp"])
        f_end = parse_date_obj(x["forecast"]["aprobacion_revp"]) or p_end
        if p_end and f_end:
            slip = (f_end - p_end).days
            if slip > 0 and x["avance_real"] < 1.0:
                delayed_items.append({
                    "activity_id": x["activity_id"],
                    "cmdic_code": x["cmdic_code"],
                    "description": x["description"],
                    "discipline": x["discipline"],
                    "type": x["type"],
                    "hh_ppto": x["hh_total_ppto"],
                    "avance_real": x["avance_real"],
                    "plan_end": x["plan"]["aprobacion_revp"],
                    "forecast_end": x["forecast"]["aprobacion_revp"],
                    "slippage_days": slip
                })
    # Sort delayed items by slippage days descending
    delayed_items.sort(key=lambda x: x["slippage_days"], reverse=True)
    
    # 5. Find most progressed items this week
    progress_gains = []
    for change in comparison["changes"]:
        if change["avance_real_diff"] > 0:
            progress_gains.append({
                "activity_id": change["activity_id"],
                "cmdic_code": change["cmdic_code"],
                "description": change["description"],
                "discipline": change["discipline"],
                "type": change["type"],
                "hh_ppto": change["hh_ppto"],
                "old_avance_real": change["old_avance_real"],
                "new_avance_real": change["new_avance_real"],
                "diff": change["avance_real_diff"]
            })
    # Sort progressed items by progress change descending
    progress_gains.sort(key=lambda x: x["diff"], reverse=True)
    
    # 6. Find tasks delayed/slipped *this week* (forecast moved out since last week)
    weekly_slips = []
    for change in comparison["changes"]:
        if change["days_slipped_this_week"] > 0:
            weekly_slips.append({
                "activity_id": change["activity_id"],
                "cmdic_code": change["cmdic_code"],
                "description": change["description"],
                "discipline": change["discipline"],
                "type": change["type"],
                "hh_ppto": change["hh_ppto"],
                "old_forecast": change["old_forecast_end"],
                "new_forecast": change["new_forecast_end"],
                "slip_days": change["days_slipped_this_week"]
            })
    weekly_slips.sort(key=lambda x: x["slip_days"], reverse=True)
    
    # 7. Consolidate Dashboard Data
    dashboard_data = {
        "kpi": {
            "semana_actual": {
                "avance_programado": actual_calcs["avance_programado"],
                "avance_real": actual_calcs["avance_real"],
                "avance_variance": actual_calcs["avance_real"] - actual_calcs["avance_programado"],
                "hh_total_ppto": actual_calcs["total_hh_ppto"],
                "hh_total_ganadas": actual_calcs["total_hh_ganadas"],
                "hh_efficiency": actual_calcs["hh_efficiency"],
                "status_counts": actual_calcs["status_counts"],
                "type_counts": actual_calcs["type_counts"],
                "total_items": actual_calcs["total_items"]
            },
            "semana_pasada": {
                "avance_programado": pasada_calcs["avance_programado"],
                "avance_real": pasada_calcs["avance_real"],
                "avance_variance": pasada_calcs["avance_real"] - pasada_calcs["avance_programado"],
                "hh_total_ppto": pasada_calcs["total_hh_ppto"],
                "hh_total_ganadas": pasada_calcs["total_hh_ganadas"],
                "hh_efficiency": pasada_calcs["hh_efficiency"],
                "status_counts": pasada_calcs["status_counts"],
                "type_counts": pasada_calcs["type_counts"],
                "total_items": pasada_calcs["total_items"]
            },
            "weekly_progress_gain": actual_calcs["avance_real"] - pasada_calcs["avance_real"],
            "weekly_earned_hh_gain": actual_calcs["total_hh_ganadas"] - pasada_calcs["total_hh_ganadas"],
            "total_slippage_count": len(delayed_items),
            "weekly_slippage_events": len(weekly_slips)
        },
        "s_curve": s_curve,
        "disciplines": actual_calcs["disciplines"],
        "delayed_items": delayed_items[:30], # Top 30 delayed items
        "progress_gains": progress_gains[:30], # Top 30 progress gains
        "weekly_slips": weekly_slips[:30], # Top 30 items that slipped this week
        "new_items": comparison["new_items"],
        "deleted_items": comparison["deleted_items"],
        "all_items": actual_items # Full raw items for search/filtering in the dashboard
    }
    
    # Save dashboard data as JS file
    output_js = "/Users/inacap/Documents/EPR/data/dashboard_data.js"
    with open(output_js, "w", encoding="utf-8") as f:
        f.write("const DASHBOARD_DATA = ")
        json.dump(dashboard_data, f, indent=2, ensure_ascii=False)
        f.write(";")
    print(f"Consolidated dashboard data saved to {output_js}.")
    
    # Save main dashboard_data.json just in case
    output_json = "/Users/inacap/Documents/EPR/data/dashboard_data.json"
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(dashboard_data, f, indent=2, ensure_ascii=False)
    print(f"JSON data saved to {output_json}.")
    print("Processing complete!")

if __name__ == "__main__":
    main()
